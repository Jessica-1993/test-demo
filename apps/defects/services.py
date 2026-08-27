import csv
import io
from pathlib import Path


class DefectImportError(ValueError):
    pass


class DefectImportService:
    COLUMN_ALIASES = {
        "缺陷编号": "defect_no",
        "编号": "defect_no",
        "缺陷标题": "title",
        "标题": "title",
        "缺陷描述": "description",
        "描述": "description",
        "复现步骤": "reproduction_steps",
        "实际结果": "actual_result",
        "预期结果": "expected_result",
        "根因": "root_cause",
        "解决方案": "resolution",
        "严重程度": "severity",
        "状态": "lifecycle_status",
        "发现版本": "detected_version_no",
        "修复版本": "fixed_version_no",
        "模块编码": "module_codes",
        "标签": "tags",
        "外部来源": "external_source",
        "外部编号": "external_id",
    }

    @classmethod
    def read_rows(cls, uploaded_file):
        suffix = Path(uploaded_file.name).suffix.lower()
        if suffix == ".csv":
            return cls._read_csv(uploaded_file.read())
        if suffix == ".xlsx":
            return cls._read_xlsx(uploaded_file)
        raise DefectImportError("仅支持 CSV 或 XLSX 文件")

    @classmethod
    def _read_csv(cls, content):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = content.decode("gb18030")
            except UnicodeDecodeError as exc:
                raise DefectImportError("CSV 文件必须使用 UTF-8 或 GB18030 编码") from exc
        return cls._normalize_rows(csv.DictReader(io.StringIO(text)))

    @classmethod
    def _read_xlsx(cls, uploaded_file):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise DefectImportError("缺少 openpyxl，无法读取 XLSX 文件") from exc
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        worksheet = workbook.active
        values = worksheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip() for value in next(values)]
        except StopIteration:
            return []
        rows = (dict(zip(headers, values_row)) for values_row in values)
        return cls._normalize_rows(rows)

    @classmethod
    def _normalize_rows(cls, rows):
        normalized = []
        for row_number, row in enumerate(rows, start=2):
            payload = {}
            for key, value in row.items():
                field = cls.COLUMN_ALIASES.get(str(key or "").strip(), str(key or "").strip())
                if field:
                    payload[field] = value.strip() if isinstance(value, str) else value
            if any(value not in (None, "") for value in payload.values()):
                normalized.append((row_number, payload))
        return normalized
